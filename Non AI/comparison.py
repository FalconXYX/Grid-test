import openpyxl
import matplotlib.pyplot as plt

def retrive(vari, row):
    wb = openpyxl.load_workbook('avg.xlsx')
    sheet = wb['1']  # Get a sheet from the workbook.
    anotherSheet = wb.active
    columbval = sheet.cell(1,2+row).value

    for x in range(2,columbval):
        vari.append(sheet.cell(x,row).value)
    print(vari)
    #wb.save("avg.xlsx")
depth = []
breath = []
retrive(depth,1)
retrive(breath,2)
ld = len(depth)
lb = len(breath)
depth = sum(depth)
breath = sum(breath)
print(depth/ld)
print(breath/lb)